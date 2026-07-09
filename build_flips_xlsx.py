#!/usr/bin/env python3
"""Build phoenix_flips.xlsx — Phoenix AZ Goodwill / FB Marketplace flip scan (July 2026)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ACCENT = "2A78D6"
HDR_FILL = PatternFill("solid", fgColor="0D366B")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="0B0B0B")
SUB_FONT = Font(size=10, color="52514E", italic=True)
HOT_FILL = PatternFill("solid", fgColor="FDECEA")
WARM_FILL = PatternFill("solid", fgColor="FFF6E5")
THIN = Border(bottom=Side(style="thin", color="E1E0D9"))

def sheet_setup(ws, title, subtitle, headers, widths):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(2, 1, subtitle).font = SUB_FONT
    for i, h in enumerate(headers, 1):
        c = ws.cell(4, i, h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.row_dimensions[4].height = 28

def fill_rows(ws, rows, heat_col=None):
    for r, row in enumerate(rows, 5):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN
        if heat_col:
            heat = row[heat_col - 1]
            if heat == "HOT":
                for c in range(1, len(row) + 1):
                    ws.cell(r, c).fill = HOT_FILL
            elif heat == "WARM":
                for c in range(1, len(row) + 1):
                    ws.cell(r, c).fill = WARM_FILL

wb = Workbook()

# ---------------- Sheet 1: Live leads ----------------
ws = wb.active
ws.title = "Live Leads Jul 2026"
sheet_setup(
    ws,
    "Phoenix Metro — Live Flip Leads (scanned July 9, 2026)",
    "Sourced from indexed Facebook Marketplace listings + estate-sale calendars. Verify each listing is still live before driving. HOT = buy-now spread.",
    ["Heat", "Item", "Area", "Ask Price", "Est. Resale", "Est. Profit", "Urgency / Desperation Signal", "Play", "Where to Look"],
    [8, 34, 14, 12, 14, 14, 38, 34, 30],
)
leads = [
    ["HOT", "West Elm Haven Sectional Sofa", "Phoenix", "$100", "$800–1,400", "$700–1,300",
     "Priced ~90% under used-market value — classic dump-it-fast pricing",
     "Grab immediately; clean, stage, photograph well. Retail $2,200+; used comps $800+.",
     "FB Marketplace › Phoenix › West Elm furniture"],
    ["HOT", "West Elm Quinn Dresser (like new)", "Mesa", "$100", "$400–600", "$300–500",
     "Like-new at 1/6 of retail = seller wants it gone",
     "Local re-list at $450 or ship-freight to Scottsdale buyers.",
     "FB Marketplace › Mesa › West Elm furniture"],
    ["HOT", "Milwaukee power tools bundle", "Phoenix/Avondale", "$150", "$300–450", "$150–300",
     "Bundle pricing under single-tool value — often job-change / quick-cash sellers",
     "Part out: individual M18 tools + batteries sell separately for 2–3x bundle price.",
     "FB Marketplace › Phoenix › Milwaukee tools"],
    ["HOT", "Mixed DeWalt/Milwaukee packages 'starting at $1'", "Phoenix metro", "$1+ (auction-style)", "varies", "high if won low",
     "Auction-style $1 starts = maximum urgency; watch for low-competition endings",
     "Bid late; verify batteries genuine (serial check) — FB tool scams exist.",
     "FB Marketplace › Phoenix › DeWalt power tools"],
    ["WARM", "West Elm Dark Gray Sectional (free delivery)", "Laveen", "$250–299", "$600–900", "$300–600",
     "Free delivery offered = seller motivated to close fast",
     "Take the free delivery, re-list staged at $700+.",
     "FB Marketplace › Phoenix › sectionals"],
    ["WARM", "Furniture moving-sale bundle", "Mesa", "$350", "$700–1,000", "$350–650",
     "Explicit 'moving sale' — July = peak Phoenix lease turnover",
     "Negotiate whole-lot price; piece out over 2 weekends.",
     "FB Marketplace › Mesa › moving sale"],
    ["WARM", "MCM West Elm Chaise, blue velvet", "Tempe", "$350–400", "$600–800", "$200–400",
     "ASU summer turnover zone — Tempe sellers clear out July/August",
     "MCM velvet photographs well; re-list to Scottsdale/Arcadia buyers.",
     "FB Marketplace › Tempe › chaise"],
    ["WARM", "Stick welder", "Buckeye", "$500", "$700–900", "$200–400",
     "Rural-edge listing, thin local buyer pool = negotiable",
     "Offer $400 cash; welders move fast closer to central Phoenix.",
     "FB Marketplace › Buckeye › welders"],
    ["WARM", "Milwaukee M18 impact wrench + XC5.0 battery", "Phoenix", "$250", "$320–380", "$70–130",
     "None visible — fair-market listing",
     "Only take at $200 or less after offer.",
     "FB Marketplace › Phoenix › Milwaukee tools"],
    ["WATCH", "DeWalt 20V MAX 4-tool combo (2 batt + charger)", "Phoenix", "$275–350", "$400–450", "$75–150",
     "None visible — margin only if talked down",
     "Offer $250; bundle with $1-start auction wins.",
     "FB Marketplace › Phoenix › DeWalt"],
    ["WATCH", "Golf clubs + cart bag, $700 OBO", "Phoenix", "$700 OBO", "part-out $900+", "$200+",
     "'OBO' flag; Phoenix golf demand is strong year-round",
     "Part out irons/driver/putter separately on eBay if brands are premium.",
     "FB Marketplace › Phoenix › golf clubs"],
    ["EVENT", "Fountain Hills estate auction — sterling silver, gold watches, vehicles", "Fountain Hills", "auction", "melt+ / comps", "varies",
     "Estate liquidation, Sat Jul 11 8:00 AM preview — no-sentiment selling",
     "Sterling flatware near spot; gold watches under melt happens at estate auctions.",
     "estatesales.org › Fountain Hills"],
    ["EVENT", "Sun City estate sale — MacKenzie-Childs, art, Subaru Outback", "Sun City", "tagged", "MC resells 2–4x", "varies",
     "Jul 11–12; Sun City estates = downsizing/passed-on, priced to clear day 2",
     "MacKenzie-Childs Courtly Check pieces resell $100–400 each on eBay.",
     "estatesales.net › Sun City"],
    ["EVENT", "Mesa estate auction — everything no-reserve", "Mesa", "no reserve", "varies", "varies",
     "NO RESERVE Sun Jul 12 = forced liquidation, best desperation format there is",
     "Show up; low crowds in July heat = thin bidding.",
     "estatesales.org › Mesa"],
    ["EVENT", "Online Phoenix estate — vintage bicycles + Swarovski collection", "Phoenix (online)", "bidding open", "varies", "varies",
     "First lots close Sun Jul 12 9:00 AM MST — online estates underprice vintage bikes",
     "Vintage road bikes (Colnago/Schwinn Paramount class) are eBay gold.",
     "estatesales.net › Phoenix online"],
]
fill_rows(ws, leads, heat_col=1)

# ---------------- Sheet 2: Goodwill playbook ----------------
ws2 = wb.create_sheet("Goodwill Flip Playbook")
sheet_setup(
    ws2,
    "Phoenix Goodwill — What to Pull & What It Sells For",
    "Buy targets for Goodwill of Central & Northern AZ retail stores, the Clearance Center bins, and shopgoodwill.com/phoenix (local pickup, 14-day window).",
    ["Category", "What Exactly", "Typical Buy", "Sold Comps", "Multiple", "Notes"],
    [22, 40, 14, 18, 10, 46],
)
playbook = [
    ["Vintage denim", "Levi's 501/517 — orange tab, Big E, 'Made in USA'", "$6–10", "$60–120", "10–20x",
     "Check the red tab and care label first; 32x32 and larger wash best."],
    ["Band/graphic tees", "Pre-2000 concert tees: Metallica, Stones, Zeppelin", "$4–8", "$80–300", "16–60x",
     "Single-stitch hems + faded prints = real. #1 $/lb item in the bins."],
    ["Vintage audio", "Pioneer, Marantz, Sansui receivers; turntables", "$25–50", "$250+", "5–10x",
     "Phoenix retiree donations feed this steadily. Test power before buying."],
    ["Video games", "CIB (complete-in-box) retro Nintendo/Sega/PS1", "$5–10", "$40+ per title", "5–10x",
     "Box + manual = 5x a loose cart. Scan lots fast with a price app."],
    ["Power tools", "DeWalt, Milwaukee, Makita, Snap-on", "$20–40", "$120+", "3–6x",
     "Goodwill underprices bare tools; batteries are the profit."],
    ["Vintage Pyrex", "Gooseberry, Friendship, Snowflake; any turquoise", "$3–8/pc", "$80–150/set", "10x+",
     "Lucky in Love pattern = $1,000–4,000. Lids triple casserole value. $0.29/lb as glassware in the bins."],
    ["Le Creuset / cast iron", "Any enameled LC; Griswold/Wagner skillets", "$10–45", "$150–300", "3–10x",
     "Check logo underside; even chipped LC sells for parts/restoration."],
    ["Athleisure", "Lululemon, Patagonia, Arc'teryx", "$6–10", "$40–90", "5–10x",
     "Rip tags tell the style/year on Lulu; Patagonia anything sells."],
    ["Furniture", "Solid wood, MCM anything, West Elm/CB2 labels", "$15–75", "$150–800", "4–10x",
     "Local FB re-list — zero shipping. Scottsdale/Arcadia buyers pay up."],
    ["Jewelry/watches", "Sterling (.925), gold-filled, running quartz watches", "$3–15", "melt to $200+", "5–20x",
     "shopgoodwill.com/phoenix jewelry lots are the sleeper; local pickup avoids shipping."],
]
fill_rows(ws2, playbook)

# ---------------- Sheet 3: Desperation signals ----------------
ws3 = wb.create_sheet("Urgency Search Strings")
sheet_setup(
    ws3,
    "Desperation & Urgency — Exact Search Strings",
    "Run these daily on FB Marketplace (radius 20 mi, sort: newest first). July–August is peak Phoenix desperation season: lease turnover, ASU move-out aftermath, heat-driven exits.",
    ["Search String", "Why It Works", "Best Time to Run"],
    [30, 52, 30],
)
signals = [
    ["\"must sell today\"", "Same-day cash need; near-zero price anchoring", "7–9 AM & 8–10 PM daily"],
    ["\"need gone asap\" / \"need gone today\"", "Space deadline (eviction, lease end, move truck booked)", "Evenings Sun–Tue"],
    ["\"moving sale\" / \"moving this weekend\"", "Hard deadline; will bundle-discount everything", "Wed–Thu before the weekend"],
    ["\"pcs\" / \"military move\"", "Luke AFB families on fixed move dates — everything must go", "End of month"],
    ["\"divorce\" / \"storage unit\"", "Emotionally motivated or fee-deadline sellers", "Any time — act fast"],
    ["\"first come first serve\"", "Seller won't hold = will take first cash offer", "Set alerts, be first"],
    ["\"obo\" + price recently dropped", "Marketplace shows price-drop history; 2+ drops = capitulation", "Filter newest, check drops"],
    ["\"free delivery\" on furniture", "Seller subsidizing the exit — motivated", "Any time"],
    ["\"husband said\" / \"wife said\"", "Household ultimatum = must move item, not maximize price", "Weekends"],
    ["\"estate\" / \"cleaning out\"", "No sentimental floor on pricing; whole-lot deals possible", "Fri–Sat mornings"],
]
fill_rows(ws3, signals)

# ---------------- Sheet 4: Locations ----------------
ws4 = wb.create_sheet("Phoenix Source Map")
sheet_setup(
    ws4,
    "Where to Source — Phoenix Metro",
    "The bins are the lowest-cost inventory in the metro. Restocks favor Monday–Tuesday after weekend donation backlog.",
    ["Source", "Location / URL", "Pricing", "Intel"],
    [32, 40, 26, 50],
)
locations = [
    ["Goodwill Clearance Center (the bins)", "51st Ave & Van Buren, Phoenix",
     "~$1.49/lb clothing · $0.89/lb hard goods · $0.29/lb glassware",
     "Rotating bins of unsold retail stock. One cart per shopper; no kids under 13. Go at rotation times; ask staff for the schedule."],
    ["Goodwill Clearance Center — East Valley", "Arizona Ave & Ray Rd, Chandler", "Same per-pound model, 9 AM–9 PM",
     "Newer location = lighter reseller competition than 51st Ave."],
    ["GCNA retail stores (100+ metro)", "goodwillaz.org store locator", "Tagged retail; weekly color-tag discounts",
     "Hit Mon/Tue after weekend donations. Scottsdale & Arcadia stores get the richest donations."],
    ["shopgoodwill.com/phoenix", "Online auction, local pickup by appointment", "Auction — snipe endings",
     "14-day pickup window. Local pickup kills shipping cost — arbitrage vs out-of-state bidders who must pay freight."],
    ["Facebook Marketplace Phoenix", "facebook.com/marketplace/phoenix", "Negotiable — use urgency strings",
     "Check seller profile age + ratings; meet at police-station exchange zones for tools/electronics."],
    ["Estate sales (estatesales.net / .org)", "Fountain Hills, Sun City, Mesa this week", "Tagged; 50% off final day",
     "Sun City = permanent downsizing pipeline. July heat thins crowds = less bidding competition."],
]
fill_rows(ws4, locations)

wb.save("/home/user/Cashbotai/phoenix_flips.xlsx")
print("saved phoenix_flips.xlsx")
